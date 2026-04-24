# -*- coding: utf-8 -*-
"""BOM 导出服务：从原始 BOM 中删去已使用库存的元件行"""
import shutil
import openpyxl
from typing import Set


def export_pruned_bom(
    original_file_path: str,
    excluded_indices: Set[int],
    output_path: str,
) -> str:
    """
    复制原始 BOM xlsx，删除 excluded_indices 对应的数据行（0-based，不含标题行）。
    重新对 No. 列（第A列）编号。
    返回输出文件路径。
    """
    shutil.copy2(original_file_path, output_path)

    wb = openpyxl.load_workbook(output_path)
    ws = wb.active

    # 数据行从 row=2 开始（row=1 为标题）
    # excluded_indices 0-based 对应 row = idx+2
    rows_to_delete = sorted(
        [idx + 2 for idx in excluded_indices],
        reverse=True,  # 从后往前删，避免行号偏移
    )
    for row_num in rows_to_delete:
        if row_num <= ws.max_row:
            ws.delete_rows(row_num)

    # 重新对 No. 列（A 列）编号
    new_no = 1
    for row in ws.iter_rows(min_row=2, max_row=ws.max_row):
        if row[0].value is not None:
            row[0].value = new_no
            new_no += 1

    wb.save(output_path)
    return output_path
